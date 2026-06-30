import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="2F4F8F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

GRADE_FILLS = {
    "⭐⭐⭐": PatternFill("solid", fgColor="FFE066"),
    "⭐⭐": PatternFill("solid", fgColor="D4EDDA"),
    "⭐": PatternFill("solid", fgColor="FFFFFF"),
    "📄": PatternFill("solid", fgColor="F8F9FA"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COL_WIDTHS = {
    "Title": 55,
    "Author / Year": 22,
    "DOI": 18,
    "Document Type": 14,
    "공정변수": 10,
    "인장물성": 10,
    "77K": 8,
    "우선순위": 10,
}


def _make_doi_url(doi: str) -> str:
    doi = doi.strip()
    if doi.startswith("http"):
        return doi
    if doi:
        return f"https://doi.org/{doi}"
    return ""


def build_excel(result_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "논문 우선순위"

    columns = list(COL_WIDTHS.keys())

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    doi_col_idx = columns.index("DOI") + 1
    grade_col_idx = columns.index("우선순위") + 1

    for row_idx, (_, row_data) in enumerate(result_df.iterrows(), start=2):
        grade = row_data["우선순위"]
        row_fill = GRADE_FILLS.get(grade, GRADE_FILLS["📄"])

        for col_idx, col_name in enumerate(columns, start=1):
            value = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_name == "Title"))
            cell.fill = row_fill

            if col_idx == doi_col_idx:
                url = _make_doi_url(str(value))
                if url:
                    cell.value = "링크"
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.value = ""
            elif col_idx == grade_col_idx:
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ("인장물성", "공정변수", "77K"):
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.value = value

        ws.row_dimensions[row_idx].height = 18

    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
